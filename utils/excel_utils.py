import openpyxl
from config import *


def read_excel(file_path=EXCEL_FILE, sheet_name=SHEET_NAME):
    # 打开 excel 文件
    print(file_path)
    workbook = openpyxl.load_workbook(file_path)  # 参数传文件路径

    # workbook = openpyxl.load_workbook("../data/测试用例.xlsx")   # 参数传文件路径

    # 选择表
    worksheet = workbook[sheet_name]

    # 空列表, 用于组装字典
    data = []
    keys = [cell.value for cell in worksheet[2]]  # 拿key行, 也就是表的第二行, 生成一个 key 的列表
    for row in worksheet.iter_rows(min_row=3, values_only=True):  # 从第三行开始拿, 只返回值
        dict_data = dict(zip(keys, row))
        # 如果读取的 is_true 字段的值是 TRUE, 则 append, 否则, 不append
        # print(dict_data["is_true"])
        if dict_data["is_true"]:
            data.append(dict_data)
        # data.append(dict_data)
    # print(data) # 打印拿到的所有数据

    # 关闭 excel 文件
    workbook.close()

    return data

# read_excel()
